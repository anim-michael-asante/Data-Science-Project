import mysql.connector as mysql
import openpyxl

mydb = mysql.connect(
    user='root',
    password='',
    host='localhost',
    database='customer_data'
)

mycursor = mydb.cursor()

WorkBook=openpyxl.load_workbook('customer_data.xlsx')

print(f"---Loading data from {WorkBook.sheetnames[0]}--- ---")

sheet = WorkBook.active

print(f'--- Reading from sheet {sheet.title} ---')

insertValues= """INSERT INTO customers (
        CustomerID, FirstName, LastName, Email, Phone, City, Country, DateJoined, AccountBalance
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""


for rows in sheet.iter_rows(min_row=2, values_only=True):
    mycursor.execute(insertValues, rows)

mydb.commit()

print(' All Excel data inserted successfully!')

